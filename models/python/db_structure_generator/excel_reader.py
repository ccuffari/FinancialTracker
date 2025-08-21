# db_structure_generator/excel_reader.py
from openpyxl import load_workbook
from typing import Dict, List, Tuple, Any
from collections import defaultdict

def read_workbook(path: str, header_row: int = 1) -> Dict[str, Dict[str, Any]]:
    """
    Legge il workbook e ritorna un dict:
    {
      "SheetName": {
         "headers": ["col1","col2",...],
         "columns": { "col1": [cell_values...], "col2": [...] },
         "has_formula": { "col1": True/False, ... },
         "first_formula_example": { "col1": "=SUM(...)", ...}
      },
      ...
    }
    """
    wb = load_workbook(path, data_only=False, read_only=True)
    sheets = {}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        headers = []
        # header row assumed header_row (1-indexed)
        header_cells = list(ws[header_row])
        headers = [ (c.value or "").strip() if c.value is not None else f"col_{i}" for i,c in enumerate(header_cells, start=1) ]
        # prepare list container for each column (rest rows)
        cols = {h: [] for h in headers}
        has_formula = {h: False for h in headers}
        first_formula_example = {h: None for h in headers}
        # iterate rows after header_row
        for r in ws.iter_rows(min_row=header_row+1, values_only=False):
            for i, cell in enumerate(r[:len(headers)]):
                col_name = headers[i]
                # if cell has formula, openpyxl stores cell.data_type == 'f' or cell.value startswith '='
                val = cell.value
                # record raw Excel cell value: if formula we capture formula text, else value
                if getattr(cell, "data_type", None) == "f" or (isinstance(val, str) and val.startswith("=")):
                    has_formula[col_name] = True
                    if first_formula_example[col_name] is None:
                        first_formula_example[col_name] = val
                    # store the formula text as marker
                    cols[col_name].append({"__formula__": val})
                else:
                    cols[col_name].append(val)
        sheets[sheetname] = {
            "headers": headers,
            "columns": cols,
            "has_formula": has_formula,
            "first_formula_example": first_formula_example,
            "raw_sheet_name": sheetname
        }
    return sheets
